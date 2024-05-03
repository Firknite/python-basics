import openpyxl

# cargamos el documento excel a trabajar
wb = openpyxl.load_workbook("planilla.xlsx")
# accedemos a los nombres de las hojas del excel
print(wb.sheetnames)
# accedemos a una hoja especifica
print(wb["Hoja 1"])
# vemos la hoja que esta activa para su lectura
hoja = wb.active
print(hoja)
# agregamos una nueva hoja
wb.create_sheet("Hoja 3")
# cambiamos titulo de una hoja
hoja3 = wb["Hoja 3"]
hoja3.title = "Nuevo titulo"
# ver cantidad de filas y columnas
print(
    hoja.max_row,
    hoja.max_column
)
# acceder a una celda especifica
celda = hoja["A1"]
celda.value = "Nombre completo"
print(celda.value)
# otra forma de acceder
celda2 = hoja.cell(row=2, column=1)
print(celda2.value)

#  recorremos todo el excel
for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        print(fila, columna, celda.value)

# obtener una columna y una fila especifica
columna = hoja["A"]
fila = hoja["1"]
print(columna)
print(fila)

# agregar mas elementos a las filas
hoja.append([1, 2, 3])
print(hoja.rows)

# eliminar filas y columnas
hoja.delete_rows(1, 1)
# hoja.delete_cols(1, 1)

# guardar todo el procesamiento en un nuevo archivo excel
wb.save("nuevo_excel.xlsx")
